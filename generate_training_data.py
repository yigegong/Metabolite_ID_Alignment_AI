import random
import json
import openpyxl
import re
from typing import List, Dict


def modify_name_with_random_delimiters(name: str, seed: int = None, remove_group: bool = False) -> str:
    """
    Modify metabolite name by randomly replacing delimiters with alternative delimiters.
    Optionally remove one random functional group to create negative pairs.

    Args:
        name: Original metabolite name
        seed: Random seed for reproducibility (optional)
        remove_group: If True, removes one random functional group (for negative pairs)

    Returns:
        Modified metabolite name with random delimiters
    """
    if seed is not None:
        random.seed(seed)

    # Alternative delimiters to use
    delimiters = [',', ' ', '-', '_', '(', ')']

    # Split by all common delimiters: comma, space, hyphen, underscore, parentheses
    parts = re.split(r'[,\s\-_()]+', name)

    # Remove empty strings from parts
    parts = [p for p in parts if p]

    # If only one part, return original name
    if len(parts) == 1:
        return name

    # If remove_group is True, remove one random functional group
    if remove_group and len(parts) > 1:
        # Randomly select one part to remove (not the first part to preserve some structure)
        remove_idx = random.randint(1, len(parts) - 1)
        parts.pop(remove_idx)

    # Randomly replace delimiters between parts
    modified_parts = [parts[0]]
    for i in range(1, len(parts)):
        delimiter = random.choice(delimiters)
        modified_parts.append(delimiter + parts[i])

    return ''.join(modified_parts)


def generate_training_pairs(
    input_file: str,
    output_file: str,
    name_column: str = 'name',
    num_positive: int = 1500,
    num_negative: int = 1500
) -> List[Dict]:
    """
    Generate positive and negative pairs from BiGG metabolites reference.

    Args:
        input_file: Path to BiGG metabolites reference file (.xlsx)
        output_file: Path to output file for training pairs (.jsonl)
        name_column: Name of the column containing metabolite names
        num_positive: Number of metabolites to use for positive pairs
        num_negative: Number of metabolites to use for negative pairs

    Returns:
        List of training data dictionaries
    """
    # Load the Excel file
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Get column headers
    headers = [cell.value for cell in ws[1]]

    # Find the name column index
    try:
        name_col_idx = headers.index(name_column)
    except ValueError:
        raise ValueError(f"Column '{name_column}' not found. Available columns: {headers}")

    # Collect all metabolite names
    all_names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        original_name = row[name_col_idx]

        if original_name is None:
            continue

        # Check if the name contains any delimiters
        if not re.search(r'[,\s\-_()]', original_name):
            continue

        all_names.append(original_name)

    # Generate training data
    training_data = []

    # Generate positive pairs from first num_positive names
    positive_names = all_names[:num_positive]
    for original_name in positive_names:
        modified_name = modify_name_with_random_delimiters(
            original_name,
            seed=hash(original_name) % (2**32),
            remove_group=False
        )

        # Create positive training pair
        training_pair = {
            "metabolite1": original_name,
            "metabolite2": modified_name,
            "label": 1  # Positive pair
        }
        training_data.append(training_pair)

    # Generate negative pairs from next num_negative names
    negative_names = all_names[num_positive:num_positive + num_negative]
    for original_name in negative_names:
        modified_name = modify_name_with_random_delimiters(
            original_name,
            seed=hash(original_name) % (2**32),
            remove_group=True  # Remove one functional group
        )

        # Create negative training pair
        training_pair = {
            "metabolite1": original_name,
            "metabolite2": modified_name,
            "label": 0  # Negative pair
        }
        training_data.append(training_pair)

    # Write to output file (JSONL format - one JSON object per line)
    with open(output_file, 'w', encoding='utf-8') as f:
        for pair in training_data:
            f.write(json.dumps(pair) + '\n')

    print(f"Generated {len(positive_names)} positive pairs from first {len(positive_names)} metabolites")
    print(f"Generated {len(negative_names)} negative pairs from next {len(negative_names)} metabolites")
    print(f"Total: {len(training_data)} training pairs")
    print(f"Output saved to {output_file}")

    return training_data


if __name__ == "__main__":
    # Configuration
    INPUT_FILE = "bigg_mets_reference.xlsx"
    OUTPUT_FILE = "training_data.jsonl"
    NAME_COLUMN = "name"  # Adjust this based on actual column name
    NUM_POSITIVE = 1500
    NUM_NEGATIVE = 1500

    # Generate training pairs (positive and negative)
    data = generate_training_pairs(
        input_file=INPUT_FILE,
        output_file=OUTPUT_FILE,
        name_column=NAME_COLUMN,
        num_positive=NUM_POSITIVE,
        num_negative=NUM_NEGATIVE
    )

    # Print some examples
    print("\nExample positive pairs:")
    for i in range(5):
        if i < len(data):
            print(f"{i+1}. {json.dumps(data[i])}")

    print("\nExample negative pairs:")
    positive_count = NUM_POSITIVE
    for i in range(5):
        idx = positive_count + i
        if idx < len(data):
            print(f"{i+1}. {json.dumps(data[idx])}")
